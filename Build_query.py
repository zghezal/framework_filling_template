import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from  Locate_Input import Locate_Input

# Convert cell coordinates to row and column numbers
def cell_to_nums(cell_coordinate):
    row_num, col_num = coordinate_to_tuple(cell_coordinate)
    return (row_num , col_num)

def find_profondeur(first_cell_nums, sheet):
    # Get the row number from first_cell_nums
    row_num = first_cell_nums[0]

    # Iterate over the cells in the row
    row = next(sheet.iter_rows(min_row=row_num+1, max_row=row_num+1))

    # Find the first and last non-empty cells in the row
    first_col = next((i for i, cell in enumerate(row) if cell.value is not None), None)
    last_col = next((i for i, cell in enumerate(reversed(row)) if cell.value is not None), None)

    if first_col is None or last_col is None:
        # If the row is empty, return None for both column numbers
        return None
    else:
        # Otherwise, return the 0-based column numbers
        return  len(row) - last_col - 1

def detect_axe_X(sheet, start_row, end_row, start_col,end_col,fill_threshold):
    # Initialize an empty DataFrame
    df = pd.DataFrame()
    inc=0
    # Iterate over the specified range of columns
    for col in range(start_col, end_col +1):
        # Initialize a list to store the cell values for this column
        col_values = []
        #Initialize a list to store the cell abscisse for this column
        col_abscisse = []
        # Initialize a counter for the number of filled cells in this column
        filled_cells = 0
        # Iterate over the specified range of rows
        for row in range(start_row, end_row + 1):
            # Get the cell value
            cell_value = sheet.cell(row=row, column=col).value
            cell_abscisse = sheet.cell(row=row, column=col).row
            # Add the cell value to the list
            col_values.append(cell_value)
            col_abscisse.append(cell_abscisse)
            # If the cell is filled, increment the counter
            if cell_value is not None and cell_value[0]!='=':
                filled_cells += 1
                

        # Calculate the fill rate for this column
        fill_rate = filled_cells / len(col_values)

        # If the fill rate is above the threshold, add the column to the DataFrame
        if fill_rate >= fill_threshold:
            inc +=1
            df['H_Level_'+str(inc)] = col_values
            df['ordinate'] = col_abscisse
    # Return the DataFrame
    return df

    
def detect_axe_Y(sheet, start_row, end_row, start_col, end_col, fill_threshold):
    # Initialize an empty DataFrame
    df = pd.DataFrame()
    inc=0

    # Iterate over the specified range of rows
    for row in range(start_row, end_row ):

        df_temp = pd.DataFrame()

        # Initialize a list to store the cell values for this row
        row_values = []
        # Initialize a list to store the cell ordinate for this row
        row_ordinate = []
        # Initialize a counter for the number of filled cells in this row
        filled_cells = 0

        # Iterate over the specified range of columns
        for col in range(start_col, end_col + 1):
        
            # Get the cell value
            cell_value = sheet.cell(row=row, column=col).value
            cell_ordinate = sheet.cell(row=row, column=col).column
            # Add the cell value to the list
            row_values.append(cell_value)
            row_ordinate.append(cell_ordinate)
            # If the cell is filled, increment the counter
            if cell_value is not None and cell_value[0]!='=':
                filled_cells += 1

        # Calculate the fill rate for this row
        fill_rate = filled_cells / len(row_values)

        # If the fill rate is above the threshold, add the row to the DataFrame
        if fill_rate >= fill_threshold:
            inc+=1
            df_temp['V_Level_'+str(inc)] = row_values
            df_temp['abscisse'] = row_ordinate
            df_temp.set_index('abscisse', inplace=True)
            df = pd.concat([df, df_temp], axis=1)
    
        fill_rate = 0

    # Return the DataFrame
    return df

def delete_formula_values(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = None
    return sheet

def duplicate_in_merged_cells(sheet):
    # Create a li1st to store the ranges of merged cells and their top left cell values
    merged_cell_ranges_and_values = [(range_, sheet[range_.start_cell.coordinate].value) for range_ in sheet.merged_cells.ranges]

    # Unmerge all the cells
    for range_, _ in merged_cell_ranges_and_values:
        sheet.unmerge_cells(str(range_))

    # Set the cell values
    for range_, top_left_cell_value in merged_cell_ranges_and_values:
        for row in sheet.iter_rows(min_row=range_.min_row, max_row=range_.max_row, min_col=range_.min_col, max_col=range_.max_col):
            for cell in row:
                cell.value = top_left_cell_value
    return sheet

def Build_query(onglet, wb, panel_asked):
    
    """Decompose un onglet en une liste de panels_aux"""    
        
    # Read the onglet in the template file
    sheet = wb[onglet]
    
    sheet = delete_formula_values(sheet)
    sheet = duplicate_in_merged_cells(sheet)
    
    # detect in sheet cell that could be a title of panel
    # title can have a font size > 12, a bold font, start with a number or a letter + a paranthesis or a point
    # extract title cell satisfying at least two of the above conditions
    from openpyxl import Workbook

    # Créer un nouveau classeur
    new_wb = Workbook()

    # Supprimer la feuille de calcul par défaut
    new_wb.remove(new_wb.active)

    # Copier la feuille de calcul dans le nouveau classeur
    new_sheet = new_wb.create_sheet(sheet.title)
    for row in sheet.iter_rows():
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value

    # Enregistrer le nouveau classeur dans un fichier
    new_wb.save('new_file.xlsx')


    title_cell = []
    # Iterate over each cell in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell is not empty
            if cell.value is not None:
                number_of_conditions = 0
                # Check if the cell has a font size > 12
                if cell.font.size is not None:
                    if cell.font.size > 12:
                        number_of_conditions += 1
                # Check if the cell has a bold font
                if cell.font.bold:
                    number_of_conditions += 1
                # Check if the cell starts with a number
                if cell.value[0].isdigit():
                    number_of_conditions += 1
                # Check if the cell starts with a letter + a paranthesis or a point
                if (cell.value[0].isalpha() and cell.value[1] == ".") or (cell.value[0].isalpha() and cell.value[1] == ")"):
                    number_of_conditions += 1
                # Check if the cell has at least two conditions
                print(number_of_conditions)
                if number_of_conditions >= 2:
                    title_cell.append(cell)
                    
    print('Title_detected', title_cell)


    panels_aux = {}
    # for each title cell, extract the panel the panel is the tranch of cell between two title cell
    for i in range(len(title_cell)):
        panel_aux=None
        # if the title cell is the last cell of the sheet, the panel is the tranch between the title cell and the end of the sheet
        if i == len(title_cell)-1:
            panel_aux = {title_cell[i].value:sheet[title_cell[i].coordinate + ':'+ get_column_letter(sheet.max_column) + str(sheet.max_row)]}
        # else the panel is the tranch between the title cell and the next title cell
        else:
            #if two title cell have consecutive row number, it is not a panel
            if title_cell[i].row == title_cell[i+1].row - 1:
                continue
            else:
                next_cell_coordinate = get_column_letter(title_cell[i].column) + str(title_cell[i].row+1)                
                
                previous_cell_coordinate = get_column_letter(title_cell[i+1].column) + str(title_cell[i+1].row-1)
                
                panel_aux = {title_cell[i].value:sheet[next_cell_coordinate:previous_cell_coordinate]}
        
        # append the panel to the list of panels_aux
        if panel_aux is not None:
            panels_aux.update(panel_aux)
    
    
    # Verifier que les panels_aux à fill demandées sont bien dans le fichier
    
    for i in panel_asked:
        # test if i is in the panel_aux keys
        if i not in panels_aux.keys():
            print('The panel',i,'is not in the',onglet,'sheet')
            print('Please check the name of the panel to fill')
            print('The panel name must be the same as the one in the',onglet,'sheet')
            print('The program will stop now')
            exit()
    
    # ne garder que les panels_aux present dans panel_asked
    panels_aux = {key: value for key, value in panels_aux.items() if key in panel_asked}
        
    panel_table ={}
    # for each panel, extract the table in the panel
    for panel_access,panel in panels_aux.items():
        # determine the two coordinate that wwould be used for the extract
        # the first coordinate is the first cell of the panel
        # the second coordinate is the last cell of the panel
        first_cell = panel[0][0].coordinate
        last_cell = panel[-1][-1].coordinate
           
        # Convert cell coordinates to row and column numbers
        first_cell_nums = cell_to_nums(first_cell)
        last_cell_nums = cell_to_nums(last_cell)

        depth=find_profondeur(first_cell_nums,sheet)

        # Read the specified range from the Excel file into a pandas DataFrame object and fill with the color of the cell 
        # Only a set of given colors have to be filled
        query_build = Locate_Input(sheet,'FFFFEC72',first_cell_nums[0],last_cell_nums[0],2,depth,panel_access)
        
        # Extraire les axes de la query
        # HORIZONTAL: abcisse
        
        # detecter les colonnes des abscisses dans le sheet
        axe_X = detect_axe_X(sheet,first_cell_nums[0],last_cell_nums[0],1,depth,0.6).reset_index(drop=True)
        
        # VERTICAL: ordonnée
        axe_Y = detect_axe_Y(sheet,first_cell_nums[0],last_cell_nums[0],1,depth,0.8)
        
        # merge query_build, axe_X and axe_Y
        query_build = query_build.merge(axe_X, how='left',left_on='ordinate',right_on='ordinate')
        query_build = query_build.merge(axe_Y,how='left', on='abscisse')

        # merge query_build and panel_table
        mapping_table = pd.read_excel('mapping_Y.xlsx',sheet_name='mapping_QIS_SA_abscisse')
        query_build = query_build.merge(mapping_table, how='left', on=['V_Level_1','V_Level_2','V_Level_3','V_Level_4'])

        # merge query_build and panel_table
        mapping_table_X = pd.read_excel('mapping_X.xlsx',sheet_name='Feuille 1')
        query_build = query_build.merge(mapping_table_X, how='left', left_on=['H_Level_2'],right_on=['Modality'])


    return query_build