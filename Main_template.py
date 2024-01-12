# this script aims to fill the template in multiple sheet of the excel file with the data from a dataframe

# Importing the libraries
import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from Build_query import Build_query
from Perform_query import Perform_query
from Write_result import Write_result

# Anlyser le fichier de template pour récupérer les noms des onglets, et le titre des templates à remplir
def Main_template(data,sheet_required,template_file,panel_asked):
    #verifier que le fichier template existe
    if not os.path.isfile(template_file):
        print('The template.xlsx file does not exist')
        print('Please check the name of the template file')
        print('The program will stop now')
        exit()
    print('The template.xlsx file exist')
    #ouvrir le fichier template
    wb = load_workbook(template_file)

    #lister les onglets du fichier template
    sheet_list = wb.sheetnames

    # Verifier que les onglets à fill demandées sont bien dans le fichier
    for i in sheet_required:
        if i not in sheet_list:
            print('The sheet',i,'is not in the template.xlsx file')
            print('Please check the name of the sheet to fill')
            print('The sheet name must be the same as the one in the template.xlsx file')
            print('The program will stop now')
            exit()

    print('The sheet to fill are in the template.xlsx file')
    
    # Loop sur les onglets à fill
    
        
    # Query Builder
    query_built = Build_query(i,wb,panel_asked)
    query_built.to_csv('query_built.csv')
        
    # Query Performing
    query_performed = Perform_query(data,query_built)
    query_performed.to_csv('query_performed.csv')
    
    # Write the result in the template
    Write_result(query_performed,template_file,i)
   

    print('finish')
    
if __name__ == "__main__":


    panel_asked=['A.1 General']
    sheet_required = ['Credit risk (SA)']
    template_file = 'template3.xlsx'
    Main_template(sheet_required,template_file,panel_asked)