# from a dataframe, loop on all line and add the result column to the coordinate columns of an excel file.
#
# Input:
# - dataframe
# - excel file
# - sheet name
# - coordinate columns
# - result column
#
# Output:
# - excel file with the result column filled
#
import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl import load_workbook

def Write_result(query_performed, wb, sheet_name,save_directly_to_excel = True):
    # Charger le fichier Excel
    book = load_workbook(wb)
    
    # Accéder à la feuille existante
    sheet = book[sheet_name]

    # Parcourir le DataFrame
    for index, row in query_performed.iterrows():
        # Obtenir la coordonnée et la valeur
        coordinate = row['coord']
        value = row['Result']
        print(coordinate+' '+value)
        # Écrire la valeur à la coordonnée spécifiée
        sheet[coordinate].value = value

    if save_directly_to_excel == True:
        # Sauvegarder le fichier Excel
        book.save(wb)
    else:
        # Sauvegarder dans un nouveau fichier Excel
        
        # Définir le nom du nouveau fichier
        wb_new = wb[:-5] + '_new.xlsx'
        sheet_new = sheet_name + '_new'
        # new book and sheet
        book.create_sheet(sheet_new)
        # Sauvegarder le fichier Excel
        book.save(wb_new)
    
    print('The result is written in the excel file')
    print('The excel file is saved')
    print('The program will stop now')
    exit()
